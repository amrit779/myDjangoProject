from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import pandas as pd
from .models import Order
from datetime import datetime

def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        workbook = openpyxl.load_workbook(excel_file)

        # getting active sheet
        excel_data = workbook.active
        
        Order.objects.all().delete()
        for column in excel_data.iter_rows(min_row=2, max_row=excel_data.max_row):
            #date = datetime.strptime(str(column[0].value), "%m/%d/%Y").strftime('%Y-%m-%d')    
            _, created = Order.objects.update_or_create(
                OrderDate = column[0].value,
                Region = str(column[1].value),
                Rep = str(column[2].value),
                Item = str(column[3].value),
                Units = str(column[4].value),
                UnitCost = str(column[5].value),
                Total = str(column[6].value),
            )
        df = pd.DataFrame(list(Order.objects.all().values()))
        #return render(request, 'index.html', {"excel_data":df})
        return render(request, 'index.html', {})

def loadData(request):
    df = pd.DataFrame(list(Order.objects.all().values()))
    #return render(request, 'index.html', {"excel_data":df})
    #return render(request, 'dashboard/home.html', {})
    #return HttpResponse(df.to_html())
    return render(request, 'dashboard/home.html', {})